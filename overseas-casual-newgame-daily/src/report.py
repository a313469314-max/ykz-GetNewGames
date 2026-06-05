from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any


REPORT_COLUMNS = [
    "canonical_company",
    "product_name",
    "store_type",
    "package",
    "app_id",
    "store_url",
    "dataeye_url",
    "scan_date",
    "first_seen_date",
    "campaign_period",
    "media_list",
    "fb_page",
    "dataeye_company_name",
    "developer_name",
    "seller_name",
    "attribution_match_type",
    "attribution_match_value",
    "attribution_confidence",
]

REPORT_TITLE = "海外休闲新品日报"


def group_by_company(products: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for product in products:
        grouped[str(product.get("canonical_company") or "Unattributed")].append(product)
    return dict(sorted(grouped.items(), key=lambda item: item[0].lower()))


def write_markdown(
    products: list[dict[str, Any]],
    report_date: str,
    path: Path | str,
    scan_label: str | None = None,
) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    grouped = group_by_company(products)
    lines = [
        f"# {REPORT_TITLE} ({report_date})",
        "",
    ]
    if scan_label:
        lines.extend([f"扫描范围：{scan_label}", ""])
    lines.extend([f"本次去重后共发现 {len(products)} 个新品，分布如下：", ""])
    for company, rows in grouped.items():
        lines.append(f"## {company} (新增 {len(rows)} 个)")
        for row in rows:
            name = row.get("product_name") or row.get("package") or row.get("app_id") or row.get("product_id") or "Unknown Product"
            link = row.get("store_url") or row.get("dataeye_url") or ""
            if link:
                lines.append(f"- {name}：{link}")
            else:
                lines.append(f"- {name}")
        lines.append("")
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return path


def _stringify(value: Any) -> str:
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    if isinstance(value, dict):
        return "; ".join(f"{k}={v}" for k, v in value.items())
    return "" if value is None else str(value)


def write_xlsx(products: list[dict[str, Any]], path: Path | str) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = REPORT_TITLE
    sheet.append(REPORT_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for row in products:
        sheet.append([_stringify(row.get(column)) for column in REPORT_COLUMNS])

    for index, column in enumerate(REPORT_COLUMNS, start=1):
        max_len = max(len(_stringify(sheet.cell(row=row_idx, column=index).value)) for row_idx in range(1, sheet.max_row + 1))
        sheet.column_dimensions[get_column_letter(index)].width = min(max(max_len + 2, 12), 60)
    sheet.freeze_panes = "A2"
    workbook.save(path)
    return path


def write_reports(
    products: list[dict[str, Any]],
    report_date: str,
    output_dir: Path | str,
    scan_label: str | None = None,
) -> tuple[Path, Path]:
    output_dir = Path(output_dir)
    markdown_path = output_dir / f"{REPORT_TITLE}_{report_date}.md"
    xlsx_path = output_dir / f"{REPORT_TITLE}_{report_date}.xlsx"
    return write_markdown(products, report_date, markdown_path, scan_label), write_xlsx(products, xlsx_path)
