from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
FIELDNAMES = ["match_type", "match_value", "canonical_company", "confidence", "source", "note"]
TARGETS = {"Homa", "Voodoo", "Rollic", "Azur", "Tapnation", "SayGames", "Supersonic"}
COMPANY_ALIAS = {"Saygame": "SayGames", "Saygames": "SayGames"}
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0 Safari/537.36"
)
REPORT_SECTION_PREFIXES = ("海外休闲新品日报", "Meta/Facebook")


def normalize_key(match_type: str, value: str) -> tuple[str, str]:
    if match_type in {"app_id", "apple_artist_id", "google_developer_id", "developer_id"}:
        return match_type, value.strip()
    return match_type, " ".join(value.strip().split()).lower()


def find_template(default_path: Path | None = None) -> Path:
    if default_path:
        return default_path
    desktop = Path.home() / "Desktop"
    matches = [path for path in desktop.glob("*.xlsx") if path.name.startswith("海外休闲新品日报")]
    if not matches:
        matches = [path for path in desktop.glob("*.xlsx") if "新品日报" in path.name]
    if not matches:
        raise FileNotFoundError("Could not find 海外休闲新品日报.xlsx on Desktop.")
    return matches[0]


def parse_template_products(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    rows = [str(row[0]) for row in workbook["Sheet1"].iter_rows(values_only=True) if row and row[0]]
    company_names = sorted(TARGETS | set(COMPANY_ALIAS), key=len, reverse=True)
    current_date = ""
    current_company = ""
    in_fb = False
    products: list[dict[str, str]] = []

    for raw_line in rows:
        line = raw_line.strip()
        if line.startswith(REPORT_SECTION_PREFIXES) and "(" in line and ")" in line:
            current_date = line.split("(", 1)[1].split(")", 1)[0]
            current_company = ""
            in_fb = True
            continue
        if line.startswith("YouTube ") or line.startswith("Youtube "):
            current_company = ""
            in_fb = False
            continue
        if not in_fb:
            continue
        if not line.startswith("- "):
            for company in company_names:
                if line.startswith(company + " "):
                    current_company = COMPANY_ALIAS.get(company, company)
                    break
            continue
        if current_company not in TARGETS:
            continue

        urls = re.findall(r"https?://[^ ]+", line.replace(chr(8230), ""))
        product_name = line[2:]
        if urls:
            product_name = product_name.split(urls[0], 1)[0].rstrip(" :")
        product_name = product_name.replace(" App - App Store", "").strip()

        for url in urls:
            parsed = urlparse(url)
            host = parsed.netloc.lower()
            query = parse_qs(parsed.query)
            product: dict[str, str] = {
                "canonical_company": current_company,
                "template_date": current_date,
                "product_name": product_name,
                "store_url": url,
            }
            if "play.google.com" in host:
                package = (query.get("id") or [""])[0]
                if package:
                    product.update({"store_type": "google_play", "package": package})
                    products.append(product)
            elif "apps.apple.com" in host or "itunes.apple.com" in host:
                match = re.search(r"/id([0-9]+)", parsed.path)
                app_id = match.group(1) if match else ""
                app_id = app_id or (query.get("id") or [""])[0]
                if app_id:
                    product.update({"store_type": "app_store", "app_id": app_id})
                    products.append(product)

    seen: set[tuple[str, str]] = set()
    deduped: list[dict[str, str]] = []
    for product in products:
        key = (product["store_type"], product.get("package") or product.get("app_id") or "")
        if key in seen:
            continue
        seen.add(key)
        deduped.append(product)
    return deduped


def load_cache(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def save_cache(path: Path, cache: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(cache, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def fetch_google_account(package: str, timeout: int = 20) -> dict[str, str]:
    url = f"https://play.google.com/store/apps/details?id={package}&hl=en&gl=US"
    response = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=timeout)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")
    result: dict[str, str] = {"package": package, "store_url": url}

    anchors = soup.select('a[href*="/store/apps/dev"], a[href*="/store/apps/developer"]')
    for anchor in anchors:
        href = anchor.get("href") or ""
        name = anchor.get_text(" ", strip=True)
        if not href or not name:
            continue
        parsed = urlparse(urljoin("https://play.google.com", href))
        developer_id = (parse_qs(parsed.query).get("id") or [""])[0]
        result.update(
            {
                "developer_name": name,
                "google_developer_id": developer_id,
                "developer_url": urljoin("https://play.google.com", href),
            }
        )
        break
    return result


def fetch_app_store_accounts(app_ids: list[str]) -> dict[str, dict[str, str]]:
    output: dict[str, dict[str, str]] = {}
    session = requests.Session()
    for index in range(0, len(app_ids), 100):
        chunk = app_ids[index : index + 100]
        response = session.get("https://itunes.apple.com/lookup", params={"id": ",".join(chunk)}, timeout=30)
        response.raise_for_status()
        payload = response.json()
        for item in payload.get("results", []):
            app_id = str(item.get("trackId") or "")
            if not app_id:
                continue
            output[app_id] = {
                "app_id": app_id,
                "product_name": str(item.get("trackName") or ""),
                "seller_name": str(item.get("sellerName") or ""),
                "developer_name": str(item.get("artistName") or ""),
                "apple_artist_id": str(item.get("artistId") or ""),
                "seller_url": str(item.get("sellerUrl") or ""),
                "bundle_id": str(item.get("bundleId") or ""),
            }
    return output


def enrich_accounts(products: list[dict[str, str]], cache_path: Path) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    cache = load_cache(cache_path)
    enriched: list[dict[str, str]] = []
    unresolved: list[dict[str, str]] = []

    google_products = [product for product in products if product["store_type"] == "google_play"]
    app_products = [product for product in products if product["store_type"] == "app_store"]

    to_fetch = [product for product in google_products if f"google:{product['package']}" not in cache]
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(fetch_google_account, product["package"]): product for product in to_fetch}
        for future in as_completed(futures):
            product = futures[future]
            key = f"google:{product['package']}"
            try:
                cache[key] = future.result()
            except Exception as exc:
                cache[key] = {"error": str(exc), "package": product["package"]}

    missing_app_ids = [product["app_id"] for product in app_products if f"apple:{product['app_id']}" not in cache]
    if missing_app_ids:
        for app_id, row in fetch_app_store_accounts(missing_app_ids).items():
            cache[f"apple:{app_id}"] = row
        for app_id in missing_app_ids:
            cache.setdefault(f"apple:{app_id}", {"error": "itunes_lookup_missing", "app_id": app_id})

    save_cache(cache_path, cache)

    for product in products:
        cache_key = f"google:{product['package']}" if product["store_type"] == "google_play" else f"apple:{product['app_id']}"
        account = cache.get(cache_key, {})
        row = dict(product)
        row.update({key: str(value) for key, value in account.items() if value not in (None, "")})
        if row.get("error") or not any(
            row.get(field) for field in ("google_developer_id", "apple_artist_id", "seller_name", "developer_name")
        ):
            unresolved.append(row)
        else:
            enriched.append(row)
    return enriched, unresolved


def candidate_mapping_rows(accounts: list[dict[str, str]]) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    candidates: dict[tuple[str, str], dict[str, Any]] = {}

    def add(row: dict[str, str], match_type: str, value: str, confidence: str) -> None:
        if not value:
            return
        key = normalize_key(match_type, value)
        entry = candidates.setdefault(
            key,
            {
                "match_type": match_type,
                "match_value": value.strip(),
                "companies": set(),
                "products": [],
                "confidence": confidence,
            },
        )
        entry["companies"].add(row["canonical_company"])
        entry["products"].append(f"{row.get('template_date')}: {row.get('product_name')}")

    for row in accounts:
        if row["store_type"] == "google_play":
            add(row, "google_developer_id", row.get("google_developer_id", ""), "high")
            add(row, "developer_name", row.get("developer_name", ""), "medium")
        else:
            add(row, "apple_artist_id", row.get("apple_artist_id", ""), "high")
            add(row, "seller_name", row.get("seller_name", ""), "high")
            add(row, "developer_name", row.get("developer_name", ""), "medium")

    mapping_rows: list[dict[str, str]] = []
    conflicts: list[dict[str, str]] = []
    for entry in candidates.values():
        companies = sorted(entry["companies"])
        products = sorted(set(entry["products"]))
        if len(companies) == 1:
            mapping_rows.append(
                {
                    "match_type": entry["match_type"],
                    "match_value": entry["match_value"],
                    "canonical_company": companies[0],
                    "confidence": entry["confidence"],
                    "source": "excel_template_account",
                    "note": "Developer account from overseas casual daily template; examples: " + " | ".join(products[:3]),
                }
            )
        else:
            conflicts.append(
                {
                    "match_type": entry["match_type"],
                    "match_value": entry["match_value"],
                    "companies": "; ".join(companies),
                    "examples": " | ".join(products[:8]),
                }
            )
    return mapping_rows, conflicts


def read_existing_mapping(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [{field: row.get(field, "") for field in FIELDNAMES} for row in csv.DictReader(handle)]


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def rewrite_company_mapping(mapping_path: Path, new_rows: list[dict[str, str]]) -> int:
    existing = read_existing_mapping(mapping_path)
    kept = [
        row
        for row in existing
        if row["source"] != "excel_template" and row["match_type"] not in {"package", "app_id", "package_name"}
    ]
    existing_keys = {normalize_key(row["match_type"], row["match_value"]) for row in kept}
    additions: list[dict[str, str]] = []
    for row in new_rows:
        key = normalize_key(row["match_type"], row["match_value"])
        if key in existing_keys:
            continue
        existing_keys.add(key)
        additions.append(row)
    write_csv(mapping_path, kept + additions, FIELDNAMES)
    return len(additions)


def main() -> int:
    parser = argparse.ArgumentParser(description="Build developer-account company mapping from the overseas casual daily Excel template.")
    parser.add_argument("--template", type=Path, help="Path to 海外休闲新品日报.xlsx. Defaults to the Desktop file.")
    parser.add_argument("--dry-run", action="store_true", help="Do not rewrite config/company_mapping.csv.")
    args = parser.parse_args()

    template_path = find_template(args.template)
    products = parse_template_products(template_path)
    accounts, unresolved = enrich_accounts(products, PROJECT_ROOT / ".cache" / "store_accounts.json")
    rows, conflicts = candidate_mapping_rows(accounts)

    write_csv(PROJECT_ROOT / "data" / "processed" / "template_store_accounts.csv", accounts, sorted({key for row in accounts for key in row}))
    write_csv(PROJECT_ROOT / "data" / "unmatched" / "template_store_account_unresolved.csv", unresolved, sorted({key for row in unresolved for key in row}) if unresolved else ["store_type", "package", "app_id", "error"])
    write_csv(PROJECT_ROOT / "data" / "unmatched" / "template_developer_account_conflicts.csv", conflicts, ["match_type", "match_value", "companies", "examples"])

    added = 0
    if not args.dry_run:
        added = rewrite_company_mapping(PROJECT_ROOT / "config" / "company_mapping.csv", rows)

    print(
        json.dumps(
            {
                "template_products": len(products),
                "resolved_accounts": len(accounts),
                "unresolved": len(unresolved),
                "candidate_mapping_rows": len(rows),
                "conflicts": len(conflicts),
                "added_to_company_mapping": added,
                "dry_run": args.dry_run,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
